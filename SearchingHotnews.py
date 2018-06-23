# -*- coding: utf-8 -*-
"""
Spyder Editor

This is a temporary script file.
"""
## BBC Website Parsing News Code
from bs4 import BeautifulSoup
import requests
import time
## 解除一些奇怪的符號
def unicodetoascii(text):
    TEXT = (text.
    		replace('\\xe2\\x80\\x99', "'").
            replace('\\xc3\\xa9', 'e').
            replace('\\xe2\\x80\\x90', '-').
            replace('\\xe2\\x80\\x91', '-').
            replace('\\xe2\\x80\\x92', '-').
            replace('\\xe2\\x80\\x93', '-').
            replace('\\xe2\\x80\\x94', '-').
            replace('\\xe2\\x80\\x94', '-').
            replace('\\xe2\\x80\\x98', "'").
            replace('\\xe2\\x80\\x9b', "'").
            replace('\\xe2\\x80\\x9c', '"').
            replace('\\xe2\\x80\\x9c', '"').
            replace('\\xe2\\x80\\x9d', '"').
            replace('\\xe2\\x80\\x9e', '"').
            replace('\\xe2\\x80\\x9f', '"').
            replace('\\xe2\\x80\\x99',"'").
            replace('\\xe2\\x80\\xa6', '...').
            replace('\\xe2\\x80\\xb2', "'").
            replace('\\xe2\\x80\\xb3', "'").
            replace('\\xe2\\x80\\xb4', "'").
            replace('\\xe2\\x80\\xb5', "'").
            replace('\\xe2\\x80\\xb6', "'").
            replace('\\xe2\\x80\\xb7', "'").
            replace('\\xe2\\x81\\xba', "+").
            replace('\\xe2\\x81\\xbb', "-").
            replace('\\xe2\\x81\\xbc', "=").
            replace('\\xe2\\x81\\xbd', "(").
            replace('\\xe2\\x81\\xbe', ")")
                 )
            
    return TEXT

def WriteBBCExcel(key):
        import openpyxl
        order=order1=order2=0
        workbook = openpyxl.load_workbook('./BBC Hexcel.xlsx')
        # 从workbook中获得所有工作表名字
        sheet = workbook.get_sheet_by_name('Hot news')
        SheetMaxrow=sheet.max_row
        ##第一行
        for col in sheet.iter_cols(min_row=SheetMaxrow+1, min_col=1, max_row=SheetMaxrow+BBCInitialnumber,max_col=1):
            for cell in col:
                cell.value =time.strftime("%Y-%m-%d", time.localtime()) 
        ##第二行
        for col in sheet.iter_cols(min_row=SheetMaxrow+1, min_col=2, max_row=SheetMaxrow+BBCInitialnumber,max_col=2):
            for cell in col:
                cell.value =BBCtheme[key]
                order=order+1
        ##第三行
        for col in sheet.iter_cols(min_row=SheetMaxrow+1, min_col=3, max_row=SheetMaxrow+BBCInitialnumber,max_col=3):
            for cell in col:
                cell.value =BBCtitles[BBCfirstpick[order1]].string.strip()
                order1=order1+1
        ##第四行
        for col in sheet.iter_cols(min_row=SheetMaxrow+1, min_col=4, max_row=SheetMaxrow+BBCInitialnumber,max_col=4):
            for cell in col:
                cell.value ="http://www.bbc.com"+BBCaddress[BBCfirstpick[order2]].get('href')          
                order2=order2+1
        print(" 目前總資料量 : ")
        print(sheet.max_row, sheet.max_column)
        workbook.save('./BBC Hexcel.xlsx')

def WriteEXPExcel(key):
        import openpyxl
        order=order1=order2=0
        workbook = openpyxl.load_workbook('./EXP Hexcel.xlsx')
        # 从workbook中获得所有工作表名字
        sheet = workbook.get_sheet_by_name('Hot news')
        SheetMaxrow=sheet.max_row
        ##第一行
        for col in sheet.iter_cols(min_row=SheetMaxrow+1, min_col=1, max_row=SheetMaxrow+EXPInitialnumber,max_col=1):
            for cell in col:
                cell.value =time.strftime("%Y-%m-%d", time.localtime()) 
        ##第二行
        for col in sheet.iter_cols(min_row=SheetMaxrow+1, min_col=2, max_row=SheetMaxrow+EXPInitialnumber,max_col=2):
            for cell in col:
                cell.value =EXPtheme[key]
                order=order+1
        ##第三行
        for col in sheet.iter_cols(min_row=SheetMaxrow+1, min_col=3, max_row=SheetMaxrow+EXPInitialnumber,max_col=3):
            for cell in col:
                cell.value = EXPID[EXPfirstpick[order1]]
                order1=order1+1
        ##第四行
        for col in sheet.iter_cols(min_row=SheetMaxrow+1, min_col=4, max_row=SheetMaxrow+EXPInitialnumber,max_col=4):
            for cell in col:
                cell.value =EXPhref[EXPfirstpick[order2]]         
                order2=order2+1
        print(" 目前總資料量 : ")
        print(sheet.max_row, sheet.max_column)
        workbook.save('./EXP Hexcel.xlsx')
        
        
        
def WriteNYExcel(key):
        import openpyxl
        order=order1=order2=order3=0
        workbook = openpyxl.load_workbook('./NY Hexcel.xlsx')
        sheet = workbook.get_sheet_by_name('Hot news')
        SheetMaxrow=sheet.max_row
        ##第一行
        for col in sheet.iter_cols(min_row=SheetMaxrow+1, min_col=1, max_row=SheetMaxrow+NYInitialnumber,max_col=1):
            for cell in col:
                cell.value =time.strftime("%Y-%m-%d", time.localtime()) 
        ##第二行
        for col in sheet.iter_cols(min_row=SheetMaxrow+1, min_col=2, max_row=SheetMaxrow+NYInitialnumber,max_col=2):
            for cell in col:
                cell.value =NYtheme[key]
                order=order+1
        ##第三行
        for col in sheet.iter_cols(min_row=SheetMaxrow+1, min_col=3, max_row=SheetMaxrow+NYInitialnumber,max_col=3):
            for cell in col:
                cell.value = NYTotal_ID[order1]
                order1=order1+1
        ##第四行
        for col in sheet.iter_cols(min_row=SheetMaxrow+1, min_col=4, max_row=SheetMaxrow+NYInitialnumber,max_col=4):
            for cell in col:
                cell.value =NYTotal_href[order2]         
                order2=order2+1
                
        ##第五行
        for col in sheet.iter_cols(min_row=SheetMaxrow+1, min_col=5, max_row=SheetMaxrow+NYInitialnumber,max_col=5):
            for cell in col:
                cell.value =NYTotal_abstract[order3]         
                order3=order3+1
        print(" 目前總資料量 : ")
        print(sheet.max_row, sheet.max_column)
        workbook.save('./NY Hexcel.xlsx')














##--------------------------Main程式-----------------------##
def BBCHotnews(key,day):
    ##  從五個類型選擇要Parse的新聞網站 key -- BBC Worlds
    global BBCtheme
    global BBCtitles
    global BBCInitialnumber
    global BBCaddress
    global BBCfirstpick
    BBCtheme=['world','science_and_environment','technology','health','business','uk']
    CurrentNumberTime=time.mktime(time.strptime(time.strftime("%a %b %d %H:%M:%S %Y", time.localtime()),"%a %b %d %H:%M:%S %Y"))
    res = requests.get('http://www.bbc.com/news/'+BBCtheme[key])
    soup = BeautifulSoup(res.text,"html.parser")
    BBCaddress=soup.select("div > a.title-link")
    BBCtitles = soup.select("div > a.title-link > h3 > span.title-link__title-text")
    timetitles=soup.select("li.mini-info-list__item > div")
    i=0
    DefaultInitialnumber=12  ## BBC前12則個符合條件的為新聞之後會為影片及其他
    BBCInitialnumber=0
    BBCfirstpick=[]
    for item in range(0,DefaultInitialnumber,1):
        i=i+1
        ##print(timetitles[i-1].string) 取得完整日期
        ##print(timetitles[i-1].get('data-seconds'))
        b=CurrentNumberTime-int(timetitles[i-1].get('data-seconds'))
        if (b/86400)<day:
            BBCInitialnumber=BBCInitialnumber+1
            BBCfirstpick.append(item)
            ##print(BBCtitles[item].string.strip())
            ##print("http://www.bbc.com"+BBCaddress[BBCfirstpick[item]].get('href'))
    if len(BBCfirstpick)==0:
        print(" No Updated News\n ")
    
    return BBCtitles,BBCaddress,BBCfirstpick 
    
    ##WriteBBCExcel(key)

def EXPHotnews(key,day,maxnumber):
    global EXPtheme
    global EXPInitialnumber
    global EXPfirstpick
    global EXPID
    global EXPhref
    EXPtheme=['world','science','finance','science-technology','health','politics','uk']
    ##---不同種類的網址不盡相桐 --##
    if key==0 or key==1 or key==5 or key==6:
        res = requests.get('https://www.express.co.uk/news/'+EXPtheme[key])
    elif key==2:
        res = requests.get('https://www.express.co.uk/'+EXPtheme[key])
    else:
        res = requests.get('https://www.express.co.uk/life-style/'+EXPtheme[key])

    soup = BeautifulSoup(res.text,"html.parser")
    titles = soup.select('ul > li > div > a')
    FirstPageNews=0
    for item in titles:
        FirstPageNews=FirstPageNews+1
    print("In the First Page, Total %d" %FirstPageNews +" News-Story")
    EXPID = {}
    EXPhref = {}
    EXPfirstpick=[]
    for j in range(0,FirstPageNews,1):
        EXPID[len(EXPID)] = titles[j].get('title')
        EXPhref[len(EXPhref)] = "https://www.express.co.uk"+titles[j].get('href') 
    ##publishtime[len(publishtime)+1]=(time.strftime("%d %B %Y"))
    ##print("%d ."%j+titles[j].get('title'))
    print("-------------------" )
    CurrentNumberTime=time.mktime(time.strptime(time.strftime("%a %b %d %H:%M:%S %Y", time.localtime()),"%a %b %d %H:%M:%S %Y"))
    EXPInitialnumber=0
    FirstPageNews=min(FirstPageNews,maxnumber)
    for k in range(0,FirstPageNews,1):
        Judgetime_res=requests.get(EXPhref[k])
        Judgetime_soup=BeautifulSoup(Judgetime_res.text,"html.parser")
        timetitles = Judgetime_soup.select('div.dates > meta')
        Thisnew_time=timetitles[0].get('content')
        time_gap=CurrentNumberTime-int(time.mktime(time.strptime(Thisnew_time,"%Y-%m-%dT%H:%M:%SZ")))
        if((time_gap/86000)<day):
            EXPInitialnumber=EXPInitialnumber+1
            EXPfirstpick.append(k)
        ##print(k,FirstPageNews)
        print(timetitles[0].get('content'))
    if len(EXPfirstpick)==0:
        print(" No Updated News\n ")
    else:
        print("Today we have %d"%EXPInitialnumber+" News-Story")    
    return EXPID,EXPhref,EXPfirstpick
    ## 寫檔紀錄
    ##WriteEXPExcel(key)


def NYHotnews(key,day):
    global NYtheme
    NYtheme=['world','science','business','technology','health','politics']
    CurrentNumberTime=time.mktime(time.strptime(time.strftime("%a %b %d %H:%M:%S %Y", time.localtime()),"%a %b %d %H:%M:%S %Y"))
    ##--New York Time 特別 分成兩個部份--##
    global NYTotal_abstract
    global NYTotal_href
    global NYTotal_ID
    NYTotal_abstract = []
    NYTotal_href = []
    NYTotal_ID = []
    global NYInitialnumber
    global NYfirstpick
    NYInitialnumber=0
    res = requests.get('https://www.nytimes.com/section/'+NYtheme[key])
    soup = BeautifulSoup(res.text,"html.parser")
    ##  先解上面四個頭條
    Four_Specialtitles=soup.select('ol.story-menu > li > article > div > h2 > a')
    Four_SpecialSummary=soup.select('ol.story-menu > li > article > div > p.summary')
    Four_SpecialTime=soup.select('ol.story-menu > li > article > div > p > span > time')
    for k in range(0,len(Four_Specialtitles),1):
        ##print(Four_Specialtitles[k].string)
        ##print(Four_SpecialSummary[k].string)
        Time_gap=CurrentNumberTime-int(Four_SpecialTime[k].get('datetime'))
        if(Time_gap/86000)<day:
            NYTotal_abstract.append(Four_SpecialSummary[k].string)
            NYTotal_href.append(Four_Specialtitles[k].get('href'))
            NYTotal_ID.append(Four_Specialtitles[k].string)
            NYInitialnumber=NYInitialnumber+1
    ##  再解下面
    address = soup.select('div.story-body > a ')
    titles = soup.select('div.story-meta > h2 ')
    abstract = soup.select('div.story-meta > p.summary')
    timeparser= soup.select('footer > time')
    i=0
    for item in range(0,int(len(titles)/2),1):
        i=i+1
        Thisnew_time=timeparser[i-1].get('content')
        Time_gap=CurrentNumberTime-int(time.mktime(time.strptime(Thisnew_time,"%Y-%m-%d")))
        if((Time_gap/86000)<day):
            NYTotal_abstract.append(abstract[i-1].string.strip())
            NYTotal_href.append(address[i-1].get('href'))
            NYTotal_ID.append(titles[i-1].string.strip())
            NYInitialnumber=NYInitialnumber+1
    return NYTotal_ID,NYTotal_href 
    ##WriteNYExcel(key)



